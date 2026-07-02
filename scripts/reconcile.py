"""
Reconcile internal_system_data.csv against external_consultant_data.xlsx using the
field mapping defined in header_mapping_table.xlsx (Header Mapping tab).

The script does NOT hardcode which external column corresponds to which canonical
field -- it reads that mapping from the workbook, so editing the mapping table
(e.g. onboarding a new plan with different consultant headers) changes the
reconciliation without touching this code.

Usage:
    python3 reconcile.py \
        --internal ../data/internal_system_data.csv \
        --external ../data/external_consultant_data.xlsx \
        --mapping ../data/header_mapping_table.xlsx \
        --output ../data/reconciliation_report.xlsx
"""
import argparse
import csv
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATE_FIELDS = {"Date of Birth", "Join Date", "Last Updated", "Created Date"}
PHONE_FIELDS = {"Phone Number"}
POSTAL_FIELDS = {"Postal Code"}
JOIN_FIELD = "External ID"


# ---------------------------------------------------------------------------
# Mapping table
# ---------------------------------------------------------------------------

def load_mapping(mapping_path):
    """Read the Header Mapping tab and return (field_order, external_header_for, merge_map)."""
    wb = load_workbook(mapping_path, data_only=True)
    ws = wb["Header Mapping"]

    field_order = []
    external_header_for = {}
    merge_notes = {}  # canonical_field -> target external header it merges into

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col = {name: i for i, name in enumerate(header)}

    for row in rows[1:]:
        if row is None or row[col["Canonical Field"]] is None:
            continue
        canonical_field = str(row[col["Canonical Field"]]).strip()
        external_header = row[col["External Consultant Header"]]
        external_header = str(external_header).strip() if external_header is not None else ""

        if canonical_field.startswith("("):
            continue  # informational row (e.g. "(no canonical equivalent)"), not a real field

        field_order.append(canonical_field)

        m = re.match(r"\(merged into (.+)\)", external_header, flags=re.IGNORECASE)
        if external_header.startswith("("):
            if m:
                merge_notes[canonical_field] = m.group(1).strip()
            continue  # no directly comparable external column (e.g. "not provided by consultant")

        external_header_for[canonical_field] = external_header

    # Resolve merge_notes (canonical field whose value gets appended to another
    # field's value before comparison, e.g. Address Line 2 -> merged into Street Address)
    merge_map = {}  # target_canonical_field -> source_canonical_field to append
    header_to_canonical = {v: k for k, v in external_header_for.items()}
    for source_field, target_external_header in merge_notes.items():
        target_field = header_to_canonical.get(target_external_header)
        if target_field:
            merge_map[target_field] = source_field

    return field_order, external_header_for, merge_map


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_internal(internal_path):
    with open(internal_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    return {row[JOIN_FIELD]: row for row in rows}, rows[0].keys() if rows else []


def load_external(external_path, join_external_header):
    wb = load_workbook(external_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    records = []
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        record = {header[i]: ("" if row[i] is None else str(row[i])) for i in range(len(header))}
        records.append(record)

    return {row[join_external_header]: row for row in records if row.get(join_external_header)}


# ---------------------------------------------------------------------------
# Normalization / comparison
# ---------------------------------------------------------------------------

def normalize_date(value):
    value = str(value or "").strip()
    if not value:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return value.lower()


def normalize_phone(value):
    return re.sub(r"\D", "", str(value or ""))


def normalize_postal(value):
    return re.sub(r"\s+", "", str(value or "")).upper()


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def normalize(field, value):
    if field in DATE_FIELDS:
        return normalize_date(value)
    if field in PHONE_FIELDS:
        return normalize_phone(value)
    if field in POSTAL_FIELDS:
        return normalize_postal(value)
    return normalize_text(value)


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

def build_internal_value(field, internal_row, merge_map):
    if internal_row is None:
        return None
    value = internal_row.get(field, "")
    source_field = merge_map.get(field)
    if source_field:
        extra = internal_row.get(source_field, "")
        if extra:
            value = f"{value}, {extra}" if value else extra
    return value


def reconcile(internal_by_key, external_by_key, comparable_fields, external_header_for, merge_map):
    all_keys = sorted(set(internal_by_key) | set(external_by_key))
    member_rows = []
    diff_rows = []

    matched = internal_only = external_only = 0
    field_mismatch_count = 0

    for key in all_keys:
        internal_row = internal_by_key.get(key)
        external_row = external_by_key.get(key)

        if internal_row and external_row:
            match_status = "Matched"
            matched += 1
        elif internal_row and not external_row:
            match_status = "Missing from External Feed"
            internal_only += 1
        else:
            match_status = "New Member (Not Yet Onboarded Internally)"
            external_only += 1

        name = ""
        if internal_row:
            name = f"{internal_row.get('First Name', '')} {internal_row.get('Last Name', '')}".strip()
        elif external_row:
            name = f"{external_row.get('First Name', '')} {external_row.get('Last Name', '')}".strip()

        member = {
            "External ID": key,
            "Name": name,
            "Plan Number": internal_row.get("Plan Number", "") if internal_row else "",
            "Plan Name": internal_row.get("Plan Name", "") if internal_row else "",
            "Internal ID": internal_row.get("Internal ID", "") if internal_row else "",
            "Match Status": match_status,
        }

        any_changed = False
        for field in comparable_fields:
            ext_header = external_header_for[field]
            internal_val = build_internal_value(field, internal_row, merge_map)
            external_val = external_row.get(ext_header, "") if external_row else None

            if internal_row and external_row:
                changed = normalize(field, internal_val) != normalize(field, external_val)
                if changed:
                    any_changed = True
                    field_mismatch_count += 1
                    diff_rows.append({
                        "External ID": key,
                        "Name": name,
                        "Difference Type": "Value Mismatch",
                        "Field": field,
                        "Internal Value": internal_val,
                        "External Value": external_val,
                    })
                changed_display = changed
            else:
                changed_display = "N/A"

            member[f"{field} (Internal)"] = internal_val if internal_val is not None else ""
            member[f"{field} (External)"] = external_val if external_val is not None else ""
            member[f"{field} Changed"] = changed_display

        member["Any Field Changed"] = any_changed if (internal_row and external_row) else "N/A"
        member_rows.append(member)

        if match_status != "Matched":
            diff_rows.append({
                "External ID": key,
                "Name": name,
                "Difference Type": match_status,
                "Field": "(entire record)",
                "Internal Value": "Present" if internal_row else "Missing",
                "External Value": "Present" if external_row else "Missing",
            })

    summary = {
        "Total Internal Members": len(internal_by_key),
        "Total External Members": len(external_by_key),
        "Matched Members": matched,
        "Missing from External Feed": internal_only,
        "New Members Not Yet Onboarded": external_only,
        "Field-Level Mismatches (Matched Members)": field_mismatch_count,
    }
    return member_rows, diff_rows, summary


# ---------------------------------------------------------------------------
# Output workbook
# ---------------------------------------------------------------------------

def style_header_row(ws, ncols, fill_color="4472C4"):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autosize(ws, rows, ncols, cap=40):
    for c in range(1, ncols + 1):
        max_len = max((len(str(r[c - 1])) for r in rows), default=10)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, cap)


def write_sheet(wb, title, columns, rows, fill_color="4472C4"):
    ws = wb.create_sheet(title)
    ws.append(columns)
    all_rows = [columns]
    for row in rows:
        values = [row.get(c, "") for c in columns]
        ws.append(values)
        all_rows.append(values)
    style_header_row(ws, len(columns), fill_color)
    autosize(ws, all_rows, len(columns))
    return ws


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--internal", default="../data/internal_system_data.csv")
    parser.add_argument("--external", default="../data/external_consultant_data.xlsx")
    parser.add_argument("--mapping", default="../data/header_mapping_table.xlsx")
    parser.add_argument("--output", default="../data/reconciliation_report.xlsx")
    args = parser.parse_args()

    field_order, external_header_for, merge_map = load_mapping(args.mapping)
    comparable_fields = [f for f in field_order if f in external_header_for and f != JOIN_FIELD]

    internal_by_key, _ = load_internal(args.internal)
    external_by_key = load_external(args.external, external_header_for[JOIN_FIELD])

    member_rows, diff_rows, summary = reconcile(
        internal_by_key, external_by_key, comparable_fields, external_header_for, merge_map
    )

    wb = Workbook()
    wb.remove(wb.active)

    member_columns = ["External ID", "Name", "Plan Number", "Plan Name", "Internal ID", "Match Status"]
    for field in comparable_fields:
        member_columns += [f"{field} (Internal)", f"{field} (External)", f"{field} Changed"]
    member_columns.append("Any Field Changed")
    write_sheet(wb, "Member Comparison", member_columns, member_rows, fill_color="4472C4")

    diff_columns = ["External ID", "Name", "Difference Type", "Field", "Internal Value", "External Value"]
    write_sheet(wb, "Differences", diff_columns, diff_rows, fill_color="C0392B")

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    for k, v in summary.items():
        ws_summary.append([k, v])
    style_header_row(ws_summary, 2, fill_color="2E7D32")
    autosize(ws_summary, [["Metric", "Value"]] + [[k, v] for k, v in summary.items()], 2)

    wb.save(args.output)

    print(f"Wrote {args.output}")
    for k, v in summary.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
